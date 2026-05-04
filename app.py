"""
app.py — CSV Email Validator v2
Fixes: EMAIL_REGEX typo, SMTP probe (STARTTLS + port 587 + MX priority sort +
definitive 550 detection), deliverability logic (no false Not Deliverable from
SMTP blocks), XLSX rebuilt only once post-validation, CSV export, UI lag reduced.
"""

import streamlit as st
import re, io, smtplib, time, random, string, ssl
import pandas as pd
from datetime import datetime, timedelta
from email_validator import validate_email as ev_validate, EmailNotValidError
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests

try:
    import dns.resolver as _dns_resolver
    DNS_AVAILABLE = True
except ImportError:
    DNS_AVAILABLE = False

# ─────────────────────────────── CONSTANTS ────────────────────────────────────
# FIX: was [a-zA-Z009.\\-]+ — digit range was broken (009 instead of 0-9)
EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", re.IGNORECASE)
TIER1 = re.compile(r"^(editor|admin|press|advert|contact)[a-z0-9._%+\-]*@", re.IGNORECASE)
TIER2 = re.compile(r"^(info|sales|hello|office|team|support|help)[a-z0-9._%+\-]*@", re.IGNORECASE)

BLOCKED_TLDS = {
    'png','jpg','jpeg','webp','gif','svg','ico','bmp','tiff','avif','mp4','mp3',
    'wav','ogg','mov','avi','webm','pdf','zip','rar','tar','gz','7z','js','css',
    'php','asp','aspx','xml','json','ts','jsx','tsx','woff','woff2','ttf','eot',
    'otf','map','exe','dmg','pkg','deb','apk'
}
PLACEHOLDER_DOMAINS = {
    'example.com','example.org','example.net','test.com','domain.com',
    'yoursite.com','yourwebsite.com','website.com','email.com','placeholder.com'
}
PLACEHOLDER_LOCALS = {
    'you','user','name','email','test','example','someone','username',
    'yourname','youremail','enter','address','sample'
}
SUPPRESS_PREFIXES = (
    'noreply','no-reply','donotreply','do-not-reply','mailer-daemon','bounce',
    'bounces','unsubscribe','notifications','notification','newsletter',
    'newsletters','postmaster','webmaster','auto-reply','autoreply','daemon',
    'robot','alerts','alert','system'
)
FREE_EMAIL_DOMAINS = {
    "gmail.com","yahoo.com","hotmail.com","outlook.com","aol.com",
    "icloud.com","protonmail.com","proton.me","zoho.com","yandex.com","mail.com",
    "gmx.com","live.com","msn.com","comcast.net","verizon.net","tutanota.com","tuta.com"
}
_DISPOSABLE_FALLBACK = {
    'tempmail.org','tempmail.net','throwawaymail.com','guerrillamailblock.com',
    'disposable-mail.com','sharklasers.com','trashmail.com','10minutemail.com',
    'maildrop.cc','tempemail.cc','getnada.com','mohmal.com','dispostable.com',
    'emailondeck.com','fakeinbox.com','grr.la','mailnesia.com','tempinbox.com',
    'tempail.com','throwaway.email','mailinator2.com','binkmail.com','bobmail.info',
    'devnullmail.com','letthemeatspam.com','reconmail.com','safetymail.info',
    'spambooger.com','spamherelots.com','20minutemail.com','30minutemail.com',
    '0-mail.com','0815.ru','0clickemail.com'
}

@st.cache_data(ttl=86400, show_spinner=False)
def fetch_disposable_domains():
    # FIX: reduced timeout from 15s → 8s to avoid blocking the whole app on startup
    try:
        r = requests.get(
            "https://raw.githubusercontent.com/disposable-email-domains/"
            "disposable-email-domains/main/disposable_email_blocklist.conf",
            timeout=8)
        if r.status_code == 200:
            online = {ln.strip().lower() for ln in r.text.splitlines()
                      if ln.strip() and not ln.startswith('#')}
            return online | _DISPOSABLE_FALLBACK
    except Exception:
        pass
    return set(_DISPOSABLE_FALLBACK)

# ─────────────────────────────── EMAIL HELPERS ────────────────────────────────
def is_valid_email(email):
    if not email: return False
    e = str(email).strip()
    if e.count('@') != 1: return False
    local, domain = e.split('@')
    lo, do = local.lower(), domain.lower()
    if not local or not domain: return False
    if local.startswith('.') or local.endswith('.') or local.startswith('-'): return False
    if len(local) > 64 or len(domain) > 255: return False
    if '.' not in domain: return False
    tld = do.rsplit('.', 1)[-1]
    if len(tld) < 2 or tld in BLOCKED_TLDS: return False
    if re.search(r'\d+x[\-\d]', do): return False
    if do in PLACEHOLDER_DOMAINS or lo in PLACEHOLDER_LOCALS: return False
    if any(lo == p or lo.startswith(p) for p in SUPPRESS_PREFIXES): return False
    if re.search(r'\d+x\d+', lo): return False
    return True

def tier_key(e):
    if TIER1.match(e): return "1"
    if TIER2.match(e): return "2"
    return "3"

def tier_short(e): return {"1": "Tier 1", "2": "Tier 2", "3": "Tier 3"}[tier_key(e)]
def sort_by_tier(emails): return sorted(emails, key=tier_key)

def confidence_score(email, val):
    if not val: return None
    s = 100
    t = tier_key(email)
    if t == "2": s -= 10
    if t == "3": s -= 25
    if not val.get("spf"):          s -= 15
    if not val.get("dmarc"):        s -= 5
    if val.get("catch_all"):        s -= 20
    if val.get("free"):             s -= 5
    st_ = val.get("status", "")
    if st_ == "Risky":              s -= 25
    if st_ == "Not Deliverable":    s -= 60
    return max(0, s)

def parse_email_cell(cell_value):
    if pd.isna(cell_value) or not str(cell_value).strip(): return []
    text = str(cell_value).strip()
    for delim in [';', ',', '|', '\n', ' * ']:
        if delim in text:
            parts = text.split(delim); break
    else:
        parts = [text]
    emails = []
    for part in parts:
        part = part.strip().strip('"').strip("'")
        found = EMAIL_REGEX.findall(part)
        if found: emails.extend(found)
        elif is_valid_email(part): emails.append(part)
    seen = set(); result = []
    for e in emails:
        el = e.lower()
        if el not in seen: seen.add(el); result.append(e)
    return result

# ─────────────────────────────── VALIDATION ENGINE ────────────────────────────
def _val_syntax(email):
    try: ev_validate(email); return True
    except EmailNotValidError: return False

def _val_mx(domain):
    try:
        recs = _dns_resolver.resolve(domain, "MX")
        # FIX: sort by preference (lowest = highest priority) — was unsorted before
        sorted_recs = sorted(recs, key=lambda r: r.preference)
        return True, [str(r.exchange) for r in sorted_recs]
    except Exception: return False, []

def _val_spf(domain):
    try:
        for rd in _dns_resolver.resolve(domain, "TXT"):
            if "v=spf1" in str(rd).lower(): return True
    except Exception: pass
    return False

def _val_dmarc(domain):
    try:
        for rd in _dns_resolver.resolve(f"_dmarc.{domain}", "TXT"):
            if "v=DMARC1" in str(rd): return True
    except Exception: pass
    return False

def _smtp_probe(email, mx_hosts, timeout=5):
    """
    FIX: replaces old _val_mailbox which had no STARTTLS, no port 587, no priority sort.

    Returns (accepted: bool, definitive: bool)
      accepted=True,  definitive=True  → RCPT TO 250 — mailbox confirmed
      accepted=False, definitive=True  → RCPT TO 550/551/553/554 — mailbox definitively absent
      accepted=False, definitive=False → could not connect / blocked / timeout (NOT a hard fail)

    Only tries the highest-priority MX host to keep timing reasonable.
    Tries port 25 plain first, then port 587 with STARTTLS.
    """
    if not mx_hosts: return False, False
    mx = mx_hosts[0].rstrip(".")

    for port, try_tls in [(25, False), (587, True)]:
        try:
            with smtplib.SMTP(timeout=timeout) as s:
                s.connect(mx, port)
                s.ehlo("validator.local")
                if try_tls:
                    try:
                        s.starttls(context=ssl.create_default_context())
                        s.ehlo("validator.local")
                    except Exception:
                        pass  # STARTTLS failed — continue on plain
                s.mail("probe@validator.local")
                code, _ = s.rcpt(email)
                if code == 250:
                    return True, True       # Accepted
                if code in (550, 551, 553, 554):
                    return False, True      # Hard rejection
                # 4xx soft bounce — not definitive, try next port
        except (smtplib.SMTPConnectError, ConnectionRefusedError, OSError):
            continue   # Port blocked — try next
        except Exception:
            continue   # Timeout or other transient — try next

    return False, False   # All attempts failed / SMTP blocked

def _catch_all_probe(domain, mx_hosts, timeout=4):
    """
    Only called when SMTP probe succeeded (smtp_definitive=True), so we know we
    can connect. Checks if the server accepts random addresses.
    """
    if not mx_hosts: return False
    mx = mx_hosts[0].rstrip(".")
    rand = ''.join(random.choices(string.ascii_lowercase + string.digits, k=22))
    fake = f"zzprobe{rand}@{domain}"
    for port, try_tls in [(25, False), (587, True)]:
        try:
            with smtplib.SMTP(timeout=timeout) as s:
                s.connect(mx, port)
                s.ehlo("validator.local")
                if try_tls:
                    try:
                        s.starttls(context=ssl.create_default_context())
                        s.ehlo("validator.local")
                    except Exception: pass
                s.mail("probe@validator.local")
                code, _ = s.rcpt(fake)
                return code == 250
        except Exception: continue
    return False

def _deliverability(syntax, mx_ok, smtp_accepted, smtp_definitive,
                    disposable, free, catch_all, spf_ok, dmarc_ok):
    """
    FIX: old logic treated SMTP timeout/block as Not Deliverable.
    Many legit custom domains (Cloudflare-routed, AWS SES, Google Workspace)
    block port 25 externally. Now only hard-fail on definitive 550 or missing MX.
    SMTP-blocked → fall back to DNS signals (SPF+DMARC) to decide.
    """
    # Hard fails — unambiguous
    if not syntax:  return "Not Deliverable", "Invalid syntax"
    if disposable:  return "Not Deliverable", "Disposable domain"
    if not mx_ok:   return "Not Deliverable", "No MX records"

    # Free providers always block external SMTP probing — trust DNS
    if free:
        return "Deliverable", "Free provider (DNS verified)"

    # Definitive SMTP rejection (550-series) — mailbox does not exist
    if smtp_definitive and not smtp_accepted:
        return "Not Deliverable", "Mailbox rejected (SMTP 550)"

    # SMTP accepted the address
    if smtp_accepted:
        if catch_all:   return "Risky", "Catch-all (mailbox accepted)"
        if not spf_ok:  return "Risky", "Verified but missing SPF"
        return "Deliverable", "Mailbox verified"

    # SMTP was blocked / timed out — NOT a hard fail
    # Use DNS signals to determine deliverability
    if catch_all:
        return "Risky", "Catch-all (SMTP blocked, unverifiable)"
    if spf_ok and dmarc_ok:
        return "Deliverable", "DNS verified (SPF + DMARC)"
    if spf_ok:
        return "Risky", "SPF present, SMTP blocked"
    return "Risky", "MX exists, SMTP blocked, no SPF"

def validate_email_full(email):
    disp   = fetch_disposable_domains()
    domain = email.split("@")[-1].lower()
    syntax = _val_syntax(email)
    mx_ok, mx_h = _val_mx(domain)   if DNS_AVAILABLE else (False, [])
    spf          = _val_spf(domain)  if DNS_AVAILABLE else False
    dmarc        = _val_dmarc(domain) if DNS_AVAILABLE else False
    disp_        = domain in disp
    free         = domain in FREE_EMAIL_DOMAINS

    smtp_accepted = smtp_definitive = catch_all = False
    if mx_ok and DNS_AVAILABLE and syntax and not disp_:
        if not free:
            smtp_accepted, smtp_definitive = _smtp_probe(email, mx_h)
            # Only probe catch-all if SMTP is actually reachable (saves time)
            if smtp_accepted or smtp_definitive:
                catch_all = _catch_all_probe(domain, mx_h)

    status, reason = _deliverability(
        syntax, mx_ok, smtp_accepted, smtp_definitive,
        disp_, free, catch_all, spf, dmarc)
    return {
        "status": status, "reason": reason, "syntax": syntax, "mx": mx_ok,
        "spf": spf, "dmarc": dmarc, "mailbox": smtp_accepted,
        "smtp_definitive": smtp_definitive, "disposable": disp_,
        "free": free, "catch_all": catch_all
    }

def validate_with_early_stop(best_email, all_emails):
    log = []
    if not best_email or not is_valid_email(best_email):
        if best_email:
            log.append((best_email, "skipped", "Invalid format"))
        for email in sort_by_tier(all_emails):
            if not is_valid_email(email) or email == best_email: continue
            val = validate_email_full(email)
            log.append((email, val["status"], val["reason"]))
            if val["status"] == "Deliverable":
                return email, val, True, log
        return best_email or (all_emails[0] if all_emails else ""), None, False, log

    val = validate_email_full(best_email)
    log.append((best_email, val["status"], val["reason"]))
    if val["status"] == "Deliverable":
        return best_email, val, False, log

    best_risky_email = best_risky_val = None
    for email in sort_by_tier(all_emails):
        if email == best_email or not is_valid_email(email): continue
        v = validate_email_full(email)
        log.append((email, v["status"], v["reason"]))
        if v["status"] == "Deliverable":
            return email, v, True, log
        if v["status"] == "Risky" and best_risky_val is None:
            best_risky_val = v; best_risky_email = email

    if best_risky_val:
        return best_risky_email, best_risky_val, True, log
    return best_email, val, False, log

# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT BUILDERS
# ══════════════════════════════════════════════════════════════════════════════
VAL_COL_ORDER = [
    "Validated Email","Status","Score","Tier","Reason",
    "SPF","DMARC","Catch-all","Fallback?","Emails Checked"
]

def _row_to_val_dict(row):
    v   = row.get("val") or {}
    em  = row.get("validated_email", "")
    cf  = row.get("confidence")
    fb  = row.get("was_fallback", False)
    st_ = v.get("status", "")
    return {
        "Validated Email": em,
        "Status":          st_ or "—",
        "Score":           cf if cf is not None else "",
        "Tier":            tier_short(em) if em else "—",
        "Reason":          v.get("reason", "—") if v else "—",
        "SPF":             ("Yes" if v.get("spf") else "No") if v else "—",
        "DMARC":           ("Yes" if v.get("dmarc") else "No") if v else "—",
        "Catch-all":       ("Yes" if v.get("catch_all") else "No") if v else "—",
        "Fallback?":       "Yes" if fb else "No",
        "Emails Checked":  len(row.get("val_log", [])),
    }

def build_csv_export(results, original_columns):
    rows = []
    for row in results:
        orig = row.get("original_row_data", {})
        r = {}
        for col in original_columns:
            val = orig.get(col, "")
            try:
                if pd.isna(val): val = ""
            except TypeError: pass
            r[col] = val
        r.update(_row_to_val_dict(row))
        rows.append(r)
    df_out = pd.DataFrame(rows, columns=list(original_columns) + VAL_COL_ORDER)
    return df_out.to_csv(index=False).encode("utf-8")

# ── XLSX builder ──────────────────────────────────────────────────────────────
def _mf(h):  return PatternFill("solid", fgColor=h)
def _fn(b=False, c="111111", s=10, n="Calibri", i=False):
    return Font(bold=b, color=c, size=s, name=n, italic=i)
def _bd():
    t = Side(style="thin", color="E5E7EB")
    return Border(left=t, right=t, top=t, bottom=t)
def _ct(): return Alignment(horizontal="center", vertical="center")
def _lt(): return Alignment(horizontal="left", vertical="center", wrap_text=False)

RF_D = _mf("F0FDF4"); RF_R = _mf("FFFBEB"); RF_B = _mf("FFF1F2"); RF_N = _mf("F9FAFB")
EF_D = _mf("DCFCE7"); EF_R = _mf("FEF3C7"); EF_B = _mf("FECACA"); EF_F = _mf("E0F2FE")
TF1  = _mf("FEF9C3"); TF2  = _mf("EEF2FF"); TF3  = _mf("F1F5F9")
CF_H = _mf("D1FAE5"); CF_M = _mf("FEF3C7"); CF_L = _mf("FEE2E2")
SF   = {"Deliverable": _mf("16A34A"), "Risky": _mf("D97706"), "Not Deliverable": _mf("DC2626")}
HDR  = _mf("111111")

def _rf(s):       return {"Deliverable": RF_D, "Risky": RF_R, "Not Deliverable": RF_B}.get(s, RF_N)
def _ef(s, fb):   return EF_F if fb else {"Deliverable": EF_D, "Risky": EF_R, "Not Deliverable": EF_B}.get(s, RF_N)
def _tf(t):       return TF1 if "1" in t else (TF2 if "2" in t else TF3)
def _cf_fill(sc): return None if sc is None else (CF_H if sc >= 75 else (CF_M if sc >= 45 else CF_L))

def _hdr(ws, r, c, v, w=None):
    cl = ws.cell(row=r, column=c, value=v)
    cl.fill = HDR; cl.font = _fn(b=True, c="FFFFFF"); cl.alignment = _ct(); cl.border = _bd()
    if w: ws.column_dimensions[get_column_letter(c)].width = w
    return cl

def _cl(ws, r, c, v, fl=None, fn_=None, al=None):
    cl = ws.cell(row=r, column=c, value=v)
    if fl:  cl.fill = fl
    if fn_: cl.font = fn_
    if al:  cl.alignment = al
    cl.border = _bd()
    return cl

def _stats_sheet(wb, name, stat_rows, title, sub=""):
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 32
    t = ws.cell(row=1, column=1, value=title)
    t.font = _fn(b=True, s=15); t.fill = _mf("F9FAFB")
    ws.merge_cells("A1:C1"); ws.row_dimensions[1].height = 28; t.alignment = _lt()
    if sub:
        s = ws.cell(row=2, column=1, value=sub)
        s.font = _fn(c="999999", s=9, i=True); ws.merge_cells("A2:C2")
    ts = ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}")
    ts.font = _fn(c="AAAAAA", s=9); ws.merge_cells("A3:C3")
    FG = {"total":"0C4A6E","deliverable":"14532D","risky":"78350F","fail":"881337",
          "fallback":"0C4A6E","none":"374151","avg":"14532D","default":"374151"}
    BG = {"total":"F0F9FF","deliverable":"F0FDF4","risky":"FFFBEB","fail":"FFF1F2",
          "fallback":"E0F2FE","none":"F9FAFB","avg":"F0FDF4","default":"F9FAFB"}
    # Use the first row with key "total" as denominator for bar charts
    denominator = max(1, next((v for _, v, k in stat_rows if k == "total"), 1))
    for i, (label, value, key) in enumerate(stat_rows, 5):
        fg = FG.get(key, FG["default"]); bg = BG.get(key, BG["default"]); fl = _mf(bg)
        _cl(ws, i, 1, label, fl, _fn(c=fg, s=10), _lt())
        _cl(ws, i, 2, value, fl, _fn(b=True, c=fg, s=11), _ct())
        ws.row_dimensions[i].height = 21
        if isinstance(value, (int, float)) and key not in ("avg",):
            pct = min(float(value) / denominator, 1.0); n = int(pct * 22)
            _cl(ws, i, 3, "█"*n + "░"*(22-n) + f"  {round(pct*100)}%",
                fl, _fn(s=9, n="Courier New", c=fg), _lt())
        else:
            _cl(ws, i, 3, "", fl)
    return ws

def build_xlsx(results, original_columns):
    wb = Workbook()
    ws = wb.active; ws.title = "Results"; ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 26

    for ci, col_name in enumerate(original_columns, 1):
        w = min(max(len(str(col_name)) * 2, 15), 40)
        _hdr(ws, 1, ci, col_name, w=w)

    val_cols_def = [
        ("Validated Email",32),("Status",16),("Score",8),("Tier",9),("Reason",24),
        ("SPF",6),("DMARC",7),("Catch-all",10),("Fallback?",10),("Emails Checked",14)
    ]
    vo = len(original_columns)
    for ci, (n, w) in enumerate(val_cols_def, vo + 1): _hdr(ws, 1, ci, n, w=w)

    for ri, row in enumerate(results, 2):
        orig_data = row.get("original_row_data", {})
        for ci, col_name in enumerate(original_columns, 1):
            val = orig_data.get(col_name, "")
            try:
                if pd.isna(val): val = ""
            except TypeError: pass
            _cl(ws, ri, ci, val, RF_N, _fn(s=9), _lt())

        v   = row.get("val") or {}
        st_ = v.get("status", "")
        fb  = row.get("was_fallback")
        em  = row.get("validated_email", "")
        cf  = row.get("confidence")
        rf  = _rf(st_); ef = _ef(st_, fb)
        vi  = vo + 1

        _cl(ws, ri, vi,   em,                              ef,           _fn(b=True, n="Courier New", s=9), _lt())
        sf_ = SF.get(st_); wc = "FFFFFF" if sf_ else "111111"
        _cl(ws, ri, vi+1, st_ or "—",                     sf_ or rf,    _fn(b=bool(sf_), c=wc, s=9), _ct())
        _cl(ws, ri, vi+2, cf if cf is not None else "—",  _cf_fill(cf) or rf, _fn(b=True, s=9), _ct())
        _cl(ws, ri, vi+3, tier_short(em) if em else "—",  _tf(tier_short(em)) if em else rf, _fn(s=9), _ct())
        _cl(ws, ri, vi+4, v.get("reason","—") if v else "—", rf, _fn(s=9), _lt())
        for c_off, key in [(5,"spf"),(6,"dmarc"),(7,"catch_all")]:
            ok = v.get(key) if v else None
            _cl(ws, ri, vi+c_off, "Yes" if ok else "No", rf,
                _fn(c="16A34A" if ok else "DC2626", s=10) if ok is not None else _fn(c="AAAAAA", s=10), _ct())
        _cl(ws, ri, vi+8, "Yes" if fb else "No", rf,
            _fn(b=bool(fb), c="0891B2" if fb else "AAAAAA", s=9), _ct())
        _cl(ws, ri, vi+9, len(row.get("val_log",[])), rf, _fn(s=9), _ct())

    # Validation Log sheet
    ws2 = wb.create_sheet("Validation Log"); ws2.freeze_panes = "A2"; ws2.row_dimensions[1].height = 26
    for ci, (n, w) in enumerate([
        ("#",6),("Domain",22),("Original Best",32),("Email Checked",32),
        ("Status",16),("Reason",22),("Result",14)], 1):
        _hdr(ws2, 1, ci, n, w)
    r2 = 2
    for ri, row in enumerate(results, 1):
        dom    = row.get("domain","")
        orig   = row.get("original_email","")
        vl     = row.get("val_log",[])
        chosen = row.get("validated_email","")
        for li, (ce, cs, cr) in enumerate(vl):
            is_f = (ce == chosen)
            rf2  = _rf(cs) if cs in ("Deliverable","Risky","Not Deliverable") else RF_N
            if is_f: rf2 = EF_D if cs=="Deliverable" else (EF_R if cs=="Risky" else EF_B)
            _cl(ws2, r2, 1, f"{ri}.{li+1}", rf2, _fn(s=9), _ct())
            _cl(ws2, r2, 2, dom if li==0 else "", rf2, _fn(s=9), _lt())
            _cl(ws2, r2, 3, orig if li==0 else "", rf2, _fn(n="Courier New",s=9,c="888888"), _lt())
            _cl(ws2, r2, 4, ce, rf2, _fn(b=is_f,n="Courier New",s=9), _lt())
            sf2 = SF.get(cs); sc = "FFFFFF" if sf2 else "111111"
            _cl(ws2, r2, 5, cs or "skipped", sf2 or rf2, _fn(b=bool(sf2),c=sc,s=9), _ct())
            _cl(ws2, r2, 6, cr, rf2, _fn(s=9), _lt())
            _cl(ws2, r2, 7, "CHOSEN" if is_f else "", EF_F if is_f else rf2,
                _fn(b=is_f, c="0891B2" if is_f else "111111", s=9), _ct())
            ws2.row_dimensions[r2].height = 15; r2 += 1

    # Stats sheet
    nt    = len(results)
    nd    = sum(1 for r in results if (r.get("val") or {}).get("status")=="Deliverable")
    nri   = sum(1 for r in results if (r.get("val") or {}).get("status")=="Risky")
    nb    = sum(1 for r in results if (r.get("val") or {}).get("status")=="Not Deliverable")
    nfb   = sum(1 for r in results if r.get("was_fallback"))
    n_emp = sum(1 for r in results if not r.get("val"))
    n_val = nt - n_emp
    tc    = sum(len(r.get("val_log",[])) for r in results)
    ac    = round(tc / n_val, 1) if n_val else 0
    confs = [r["confidence"] for r in results if r.get("confidence") is not None]
    avgc  = round(sum(confs)/len(confs), 1) if confs else "—"
    _stats_sheet(wb, "Stats", [
        ("Total rows in CSV",            nt,    "total"),
        ("Rows with no email (retained)", n_emp,"none"),
        ("Rows validated",               n_val, "total"),
        ("Deliverable",                  nd,    "deliverable"),
        ("Risky",                        nri,   "risky"),
        ("Not Deliverable",              nb,    "fail"),
        ("Fallback emails used",         nfb,   "fallback"),
        ("Total emails checked",         tc,    "total"),
        ("Avg checks per row",           ac,    "avg"),
        ("Avg confidence score",         avgc,  "avg"),
    ], "CSV Email Validator — Results",
       f"{nt} total · {n_val} validated · {n_emp} empty retained")

    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT APP
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="CSV Email Validator", page_icon="✅",
                   layout="wide", initial_sidebar_state="expanded")
ACCENT = "#16a34a"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');
*,html,body,[class*="css"]{{font-family:'Inter',system-ui,sans-serif!important;}}
#MainMenu,footer,header{{visibility:hidden;}}
.block-container{{padding:1.2rem 2rem 4rem!important;max-width:100%!important;background:#f6f5f2!important;}}
[data-testid="stSidebar"]{{background:#111!important;}}
[data-testid="stSidebar"] *{{color:#ccc!important;}}
[data-testid="stSidebar"] .stDownloadButton>button{{background:{ACCENT}!important;border:none!important;color:#fff!important;border-radius:8px!important;font-size:12px!important;font-weight:700!important;width:100%!important;margin-bottom:6px!important;}}
[data-testid="stSidebar"] .stDownloadButton>button:hover{{opacity:.88!important;}}
.mh-ph{{display:flex;align-items:center;gap:12px;padding:14px 20px;background:#fff;border:1px solid #e8e8e4;border-radius:12px;margin-bottom:16px;}}
.mh-pi{{width:38px;height:38px;border-radius:10px;background:{ACCENT};display:flex;align-items:center;justify-content:center;font-size:18px;color:#fff;flex-shrink:0;}}
.mh-pt{{font-size:17px;font-weight:800;color:#111;letter-spacing:-.4px;}}
.mh-ps{{font-size:11px;color:#aaa;margin-top:1px;}}
.mh-sec{{font-size:9.5px;font-weight:700;letter-spacing:1.3px;text-transform:uppercase;color:#c0bfbb;display:block;margin-bottom:6px;}}
.stButton>button{{font-family:'Inter',sans-serif!important;font-weight:600!important;border-radius:8px!important;font-size:12.5px!important;height:36px!important;transition:all .13s ease!important;}}
.stButton>button[kind="primary"]{{background:{ACCENT}!important;border:2px solid {ACCENT}!important;color:#fff!important;box-shadow:0 1px 3px rgba(0,0,0,.15)!important;}}
.stButton>button[kind="primary"]:hover{{opacity:.88!important;transform:translateY(-1px)!important;box-shadow:0 4px 12px rgba(0,0,0,.2)!important;}}
.stButton>button[kind="primary"]:disabled{{background:#e6e6e4!important;border-color:#e6e6e4!important;color:#bbb!important;box-shadow:none!important;transform:none!important;opacity:1!important;}}
.stButton>button[kind="secondary"]{{background:#fff!important;border:1.5px solid #ddd!important;color:#555!important;}}
.stButton>button[kind="secondary"]:hover{{border-color:{ACCENT}!important;color:{ACCENT}!important;}}
.mh-big .stButton>button{{height:46px!important;font-size:14px!important;font-weight:800!important;}}
.stDownloadButton>button{{font-family:'Inter',sans-serif!important;font-weight:600!important;border-radius:8px!important;font-size:12.5px!important;height:36px!important;background:{ACCENT}!important;border:none!important;color:#fff!important;}}
[data-testid="stFileUploader"]{{background:#fff!important;border:1.5px dashed #e4e4e0!important;border-radius:8px!important;}}
.vp{{height:4px;background:#f0f0ee;border-radius:99px;overflow:hidden;margin:6px 0;}}
.vf{{height:100%;border-radius:99px;background:{ACCENT};transition:width .35s;}}
.mh-log{{background:#18181b;border-radius:8px;padding:10px 12px;font-family:'JetBrains Mono',monospace;font-size:10.5px;line-height:1.8;max-height:200px;overflow-y:auto;margin-top:6px;}}
.mh-log::-webkit-scrollbar{{width:4px;}}
.mh-log::-webkit-scrollbar-thumb{{background:#3f3f46;border-radius:2px;}}
.lr{{color:#fff;font-weight:700;border-top:1px solid #27272a;margin-top:4px;padding-top:4px;}}
.lr:first-child{{border-top:none;margin-top:0;padding-top:0;}}
.lo{{color:#4ade80;font-weight:600;}}
.lf{{color:#f87171;}}
.ls{{color:#fb923c;}}
.li{{color:#3f3f46;}}
.lx{{color:#22d3ee;font-weight:700;}}
.mh-info{{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:8px 13px;font-size:12px;color:#15803d;font-weight:600;margin:4px 0;}}
.mh-warn{{background:#fff1f2;border:1px solid #fecdd3;border-radius:8px;padding:8px 13px;font-size:12px;color:#be123c;font-weight:600;margin:4px 0;}}
.cp{{background:#fafaf8;border:1px solid #e8e8e4;border-radius:8px;padding:10px 14px;margin:6px 0;font-size:11.5px;}}
.cp-l{{font-size:9.5px;font-weight:700;color:#999;text-transform:uppercase;letter-spacing:1px;margin-bottom:4px;}}
.cp-v{{font-family:'JetBrains Mono',monospace;font-size:11px;color:#333;line-height:1.6;}}
hr{{border-color:#eee!important;margin:10px 0!important;}}
</style>""", unsafe_allow_html=True)

# ── Session state init ─────────────────────────────────────────────────────────
_SS_DEFAULTS = {
    "cv_results": [], "cv_running": False, "cv_idx": 0, "cv_log": [],
    "cv_queue": [], "cv_original_cols": [], "cv_start_time": 0.0,
    "cv_xlsx": None, "cv_csv": None,
}
for k, v in _SS_DEFAULTS.items():
    if k not in st.session_state: st.session_state[k] = v

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div style="font-size:17px;font-weight:800;color:#fff;letter-spacing:-.3px;margin-bottom:4px">CSV Validator</div>', unsafe_allow_html=True)
    st.markdown('<div style="font-size:10px;color:#555;margin-bottom:16px">early-stop · DNS + SMTP · STARTTLS</div>', unsafe_allow_html=True)
    st.divider()

    res     = st.session_state.cv_results
    running = st.session_state.cv_running

    if res and not running:
        # FIX: XLSX/CSV built ONCE after run completes, not on every rerun
        orig_cols = st.session_state.cv_original_cols
        if st.session_state.cv_xlsx is None:
            with st.spinner("Building exports…"):
                st.session_state.cv_xlsx = build_xlsx(res, orig_cols)
                st.session_state.cv_csv  = build_csv_export(res, orig_cols)
        ts = datetime.now().strftime('%Y%m%d_%H%M')
        st.download_button(
            "📥 Export .xlsx", st.session_state.cv_xlsx,
            f"validated_{ts}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="sx_xlsx", use_container_width=True)
        st.download_button(
            "📥 Export .csv", st.session_state.cv_csv,
            f"validated_{ts}.csv", "text/csv",
            key="sx_csv", use_container_width=True)
        st.divider()
        nd_  = sum(1 for r in res if (r.get("val") or {}).get("status")=="Deliverable")
        nri_ = sum(1 for r in res if (r.get("val") or {}).get("status")=="Risky")
        nb_  = sum(1 for r in res if (r.get("val") or {}).get("status")=="Not Deliverable")
        nfb_ = sum(1 for r in res if r.get("was_fallback"))
        st.markdown(
            f'<div style="font-size:11px;color:#666;line-height:2.4">'
            f'Total: <strong style="color:#fff">{len(res)}</strong><br>'
            f'Deliverable: <strong style="color:#4ade80">{nd_}</strong><br>'
            f'Risky: <strong style="color:#fb923c">{nri_}</strong><br>'
            f'Failed: <strong style="color:#f87171">{nb_}</strong><br>'
            f'Fallbacks used: <strong style="color:#22d3ee">{nfb_}</strong></div>',
            unsafe_allow_html=True)
    elif running:
        done = len(st.session_state.cv_results)
        tot  = len(st.session_state.cv_queue)
        st.markdown(f'<div style="color:#fff;font-size:12px">Validating… {done}/{tot}</div>', unsafe_allow_html=True)

    st.divider()
    st.markdown(
        '<div style="font-size:9px;color:#333;line-height:1.9">'
        'Upload CSV → pick columns<br>'
        'Best Email validated first<br>'
        'All Emails = fallback pool<br>'
        'Empty rows kept in export<br>'
        'Stops on first Deliverable<br><br>'
        '<strong style="color:#555">v2 logic:</strong><br>'
        'Free providers → DNS trusted<br>'
        'SMTP 550 → Not Deliverable<br>'
        'SMTP blocked → DNS fallback<br>'
        'STARTTLS + port 587 probing<br>'
        'MX sorted by priority</div>',
        unsafe_allow_html=True)

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown(
    f'<div class="mh-ph"><div class="mh-pi">✅</div><div>'
    f'<div class="mh-pt">CSV Email Validator</div>'
    f'<div class="mh-ps">upload CSV · early-stop fallback · DNS + SMTP (STARTTLS) · XLSX &amp; CSV export</div>'
    f'</div></div>', unsafe_allow_html=True)

# ── Upload ─────────────────────────────────────────────────────────────────────
st.markdown('<span class="mh-sec">Upload CSV</span>', unsafe_allow_html=True)
uploaded = st.file_uploader("Choose a CSV file", type=["csv"], key="csv_up")
df = None
if uploaded:
    try:
        df = pd.read_csv(uploaded)
        cols = list(df.columns)
        st.markdown(f'<div class="mh-info">Loaded <strong>{len(df)}</strong> rows · <strong>{len(cols)}</strong> columns</div>', unsafe_allow_html=True)
        st.caption("Preview (first 5 rows)")
        st.dataframe(df.head(5), use_container_width=True, hide_index=True, height=160)
    except Exception as e:
        st.error(f"Failed to parse CSV: {e}")

# ── Column selection + controls ────────────────────────────────────────────────
if df is not None:
    cols = list(df.columns)
    bh = ["best email","best_email","primary email","email","mail","contact email","contact_email"]
    ah = ["all emails","all_emails","emails","other emails","other_emails",
          "additional emails","additional_emails","alt emails","fallback"]
    dh = ["domain","website","site","url","company"]
    db = next((c for c in cols if any(h in c.lower() for h in bh)), cols[0])
    da = next((c for c in cols if any(h in c.lower() for h in ah)), None)
    dd = next((c for c in cols if any(h in c.lower() for h in dh)), None)

    st.divider()
    st.markdown('<span class="mh-sec">Column Mapping</span>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3, gap="large")
    with c1:
        st.markdown('<div style="font-size:12px;font-weight:700;color:#111;margin-bottom:4px">Best Email *</div>', unsafe_allow_html=True)
        best_col = st.selectbox("b", cols, index=cols.index(db), key="s_b", label_visibility="collapsed")
    with c2:
        st.markdown('<div style="font-size:12px;font-weight:700;color:#111;margin-bottom:4px">All Emails (fallback pool)</div>', unsafe_allow_html=True)
        all_col = st.selectbox("a", ["— None —"] + cols,
                               index=(cols.index(da)+1 if da else 0), key="s_a", label_visibility="collapsed")
    with c3:
        st.markdown('<div style="font-size:12px;font-weight:700;color:#111;margin-bottom:4px">Domain (optional)</div>', unsafe_allow_html=True)
        dom_col = st.selectbox("d", ["— Auto —"] + cols,
                               index=(cols.index(dd)+1 if dd else 0), key="s_d", label_visibility="collapsed")

    st.markdown(
        f'<div class="cp"><div class="cp-l">Best Email column</div><div class="cp-v">' +
        "<br>".join(str(v)[:60] for v in df[best_col].head(3).values) +
        '</div></div>', unsafe_allow_html=True)
    if all_col != "— None —":
        st.markdown(
            f'<div class="cp"><div class="cp-l">All Emails column</div><div class="cp-v">' +
            "<br>".join(str(v)[:80] for v in df[all_col].head(3).values) +
            '</div></div>', unsafe_allow_html=True)

    # FIX: Only rebuild queue when NOT running — during validation use cached queue
    # This prevents re-iterating the entire CSV on every rerun (was O(n) every cycle)
    running = st.session_state.cv_running
    if not running:
        queue = []
        for i, row in df.iterrows():
            br  = str(row[best_col]).strip() if pd.notna(row[best_col]) else ""
            be  = br if is_valid_email(br) else ""
            ar  = str(row[all_col]).strip() if (all_col != "— None —" and pd.notna(row.get(all_col))) else ""
            ae  = [e for e in (parse_email_cell(ar) if ar else []) if e.lower() != be.lower()]
            has = bool(be or ae)
            if not has:
                dom = f"row_{i+1}"
            elif dom_col != "— Auto —" and pd.notna(row.get(dom_col)):
                dom = str(row[dom_col]).strip()
            elif be:
                dom = be.split("@")[-1]
            else:
                dom = ae[0].split("@")[-1]
            queue.append({
                "row_idx": i+1, "domain": dom,
                "original_email": be, "all_emails": ae,
                "has_emails": has, "original_row_data": row.to_dict()
            })
    else:
        queue = st.session_state.cv_queue  # Reuse cached queue during validation

    nv      = sum(1 for q in queue if q["has_emails"])
    n_empty = len(queue) - nv
    st.divider()
    if nv:
        e_txt = f" · <strong>{n_empty}</strong> empty rows retained in export" if n_empty else ""
        st.markdown(f'<div class="mh-info">Validatable: <strong>{nv}</strong> rows{e_txt}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="mh-warn">No valid emails found. Check column mapping.</div>', unsafe_allow_html=True)

    vc1, vc2, vc3 = st.columns([3, 1, 2])
    with vc1:
        st.markdown('<div class="mh-big">', unsafe_allow_html=True)
        if not running:
            if st.button(f"▶  Validate {nv} row(s)", type="primary",
                         use_container_width=True, disabled=not nv, key="cv_go"):
                st.session_state.cv_results       = []
                st.session_state.cv_idx           = 0
                st.session_state.cv_log           = []
                st.session_state.cv_running       = True
                st.session_state.cv_queue         = queue
                st.session_state.cv_original_cols = list(df.columns)
                st.session_state.cv_start_time    = time.time()
                st.session_state.cv_xlsx          = None   # Invalidate cached exports
                st.session_state.cv_csv           = None
                st.rerun()
        else:
            if st.button("⏹  Stop", type="secondary", use_container_width=True, key="cv_stop"):
                st.session_state.cv_running = False; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    with vc2:
        if st.session_state.cv_results and not running:
            if st.button("Clear", type="secondary", use_container_width=True, key="cv_clr"):
                for k in ("cv_results","cv_log","cv_queue"):
                    st.session_state[k] = []
                st.session_state.cv_xlsx = None
                st.session_state.cv_csv  = None
                st.rerun()
    with vc3:
        st.markdown('<div style="font-size:10.5px;color:#aaa;padding-top:12px">~3-10s/email · stops on first Deliverable</div>', unsafe_allow_html=True)

    # ── Live Progress ──────────────────────────────────────────────────────────
    res = st.session_state.cv_results
    cq  = st.session_state.cv_queue
    ci  = st.session_state.cv_idx

    if running and cq:
        nv_total    = sum(1 for q in cq if q["has_emails"])
        valid_done  = sum(1 for r in res if r.get("has_emails"))
        nd_l  = sum(1 for r in res if (r.get("val") or {}).get("status")=="Deliverable")
        nri_l = sum(1 for r in res if (r.get("val") or {}).get("status")=="Risky")
        nb_l  = sum(1 for r in res if (r.get("val") or {}).get("status")=="Not Deliverable")
        elapsed   = time.time() - st.session_state.cv_start_time
        speed     = valid_done / elapsed if elapsed > 0 else 0
        remaining = nv_total - valid_done
        eta_sec   = remaining / speed if speed > 0 else 0
        eta_str   = str(timedelta(seconds=int(eta_sec))) if remaining > 0 else "done"
        pct       = round(valid_done / nv_total * 100, 1) if nv_total else 0
        cur       = cq[ci] if ci < len(cq) else None
        ce        = ((cur.get("original_email") or
                     (cur["all_emails"][0] if cur and cur.get("all_emails") else "—"))
                    if cur else "—")
        st.markdown(
            f'<div style="font-size:12px;font-weight:700;color:#111;margin:6px 0 2px">'
            f'Validating {valid_done}/{nv_total} — '
            f'<code style="color:{ACCENT}">{ce[:48]}</code></div>'
            f'<div class="vp"><div class="vf" style="width:{pct}%"></div></div>'
            f'<div style="font-size:20px;font-weight:800;color:{ACCENT};text-align:right;margin-top:-4px">{pct}%</div>',
            unsafe_allow_html=True)
        st.markdown(
            f'<div style="font-size:11px;color:#666;display:flex;gap:18px;flex-wrap:wrap;margin:8px 0">'
            f'<span>✅ <strong style="color:#16a34a">{nd_l}</strong></span>'
            f'<span>⚠️ <strong style="color:#d97706">{nri_l}</strong></span>'
            f'<span>❌ <strong style="color:#dc2626">{nb_l}</strong></span>'
            f'<span>Speed: <strong>{speed:.2f}/s</strong></span>'
            f'<span>ETA: <strong>{eta_str}</strong></span></div>',
            unsafe_allow_html=True)

    # ── Log terminal ───────────────────────────────────────────────────────────
    ll = st.session_state.cv_log
    if ll:
        h = ""
        for kind, text in ll[-60:]:
            if   kind == "row":  h += f'<div class="lr">[ {text} ]</div>'
            elif kind == "try":  h += f'<div class="li">  ↳ {text}</div>'
            elif kind == "ok":   h += f'<div class="lo">  ✓ {text}</div>'
            elif kind == "fail": h += f'<div class="lf">  ✗ {text}</div>'
            elif kind == "skip": h += f'<div class="ls">  ⤳ {text}</div>'
            elif kind == "stop": h += f'<div class="lx">  ⏹ {text}</div>'
        st.markdown(f'<div class="mh-log">{h}</div>', unsafe_allow_html=True)

    # ── Results table ──────────────────────────────────────────────────────────
    # FIX: during validation show only last 10 rows (not full styled DF on every rerun)
    if res:
        display_res = res[-10:] if running else res
        rows_tbl = []
        for r in display_res:
            v    = r.get("val") or {}
            s    = v.get("status","")
            em   = r.get("validated_email","")
            orig = r.get("original_email","")
            fb   = r.get("was_fallback")
            cf   = r.get("confidence")
            if not r.get("has_emails"):
                rows_tbl.append({"#":r["row_idx"],"Domain":r["domain"],"Validated":"—",
                                  "Status":"Skipped","Tier":"—","Score":"—","Reason":"—","FB":""})
                continue
            rows_tbl.append({
                "#":        r["row_idx"],
                "Domain":   r["domain"],
                "Validated": f"{em} ↩" if fb and orig != em else em,
                "Status":   s or "Pending",
                "Tier":     tier_short(em) if em else "—",
                "Score":    cf if cf is not None else "—",
                "Reason":   v.get("reason","—") if v else "—",
                "FB":       "Yes" if fb else "",
            })

        dr = pd.DataFrame(rows_tbl)
        status_col_idx = list(dr.columns).index("Status")

        def color_status(row):
            styles = [""] * len(row)
            if   row["Status"] == "Deliverable":     styles[status_col_idx] = "background-color:#d4edda;color:#155724"
            elif row["Status"] == "Risky":           styles[status_col_idx] = "background-color:#fff3cd;color:#856404"
            elif row["Status"] == "Not Deliverable": styles[status_col_idx] = "background-color:#f8d7da;color:#721c24"
            elif row["Status"] == "Skipped":         styles[status_col_idx] = "background-color:#f9fafb;color:#999"
            return styles

        caption = (f"Last 10 results (live) — {len(res)} total so far" if running
                   else f"**{len(res)}** rows complete  |  ↩ = fallback email used")
        st.caption(caption)
        st.dataframe(
            dr.style.apply(color_status, axis=1),
            use_container_width=True, hide_index=True,
            height=min(520, 44 + max(len(dr), 1) * 36),
            column_config={
                "#":         st.column_config.NumberColumn("#", width=45),
                "Domain":    st.column_config.TextColumn("Domain", width=140),
                "Validated": st.column_config.TextColumn("Validated Email", width=230),
                "Status":    st.column_config.TextColumn("Status", width=145),
                "Tier":      st.column_config.TextColumn("Tier", width=65),
                "Score":     st.column_config.NumberColumn("Score", width=55),
                "Reason":    st.column_config.TextColumn("Reason", width=200),
                "FB":        st.column_config.TextColumn("FB", width=38),
            })

    # ── VALIDATION ENGINE (runs at end of each rerun) ─────────────────────────
    if st.session_state.cv_running:
        q   = st.session_state.cv_queue
        idx = st.session_state.cv_idx

        if idx >= len(q):
            st.session_state.cv_running = False; st.rerun()
        else:
            item      = q[idx]
            rn        = item["row_idx"]
            dom       = item["domain"]
            orig_data = item.get("original_row_data", {})

            if not item.get("has_emails"):
                # Batch all consecutive empty rows in one rerun to reduce rerun count
                next_idx = idx
                while next_idx < len(q) and not q[next_idx].get("has_emails"):
                    ni = q[next_idx]
                    st.session_state.cv_results.append({
                        "row_idx": ni["row_idx"], "domain": ni["domain"],
                        "original_email": "", "validated_email": "", "all_emails": [],
                        "val": None, "was_fallback": False, "confidence": None,
                        "val_log": [], "original_row_data": ni.get("original_row_data", {}),
                        "has_emails": False
                    })
                    next_idx += 1
                st.session_state.cv_idx = next_idx
                if next_idx >= len(q): st.session_state.cv_running = False
                st.rerun()
            else:
                best = item["original_email"]
                ae   = item["all_emails"]
                st.session_state.cv_log.append(("row",  f"Row {rn} — {dom}"))
                st.session_state.cv_log.append(("try",  f"Best: {best or '(none)'}"))

                val_em, val_res, was_fb, vlog = validate_with_early_stop(best, ae)

                for ce, cs, cr in vlog:
                    if cs == "Deliverable":
                        st.session_state.cv_log.append(("ok",   f"{ce} — DELIVERABLE"))
                        st.session_state.cv_log.append(("stop", f"Stopping — deliverable found"))
                    elif cs == "Risky":
                        st.session_state.cv_log.append(("try",  f"{ce} — Risky ({cr})"))
                    elif cs == "Not Deliverable":
                        st.session_state.cv_log.append(("fail", f"{ce} — {cr}"))
                    else:
                        st.session_state.cv_log.append(("skip", f"{ce} — {cr}"))
                if was_fb:
                    st.session_state.cv_log.append(("ok", f"Fallback used: {best} → {val_em}"))

                # FIX: trim log in session state to avoid unbounded memory growth
                if len(st.session_state.cv_log) > 500:
                    st.session_state.cv_log = st.session_state.cv_log[-300:]

                cf = confidence_score(val_em, val_res) if val_res else None
                st.session_state.cv_results.append({
                    "row_idx": rn, "domain": dom,
                    "original_email": best, "validated_email": val_em,
                    "all_emails": ae, "val": val_res,
                    "was_fallback": was_fb, "confidence": cf,
                    "val_log": vlog, "original_row_data": orig_data,
                    "has_emails": True
                })
                st.session_state.cv_idx = idx + 1
                if st.session_state.cv_idx >= len(q):
                    st.session_state.cv_running = False
                st.rerun()

# ── Empty state ────────────────────────────────────────────────────────────────
if df is None and not st.session_state.cv_results:
    st.markdown("""
    <div style="text-align:center;padding:60px 0">
        <div style="font-size:48px;opacity:.08;margin-bottom:16px">✅</div>
        <div style="font-size:18px;font-weight:800;color:#111;margin-bottom:10px">Upload a CSV to start</div>
        <div style="font-size:12.5px;color:#aaa;line-height:2.1;max-width:460px;margin:0 auto">
            Your CSV needs at least one email column.<br>
            Optionally a second column with additional fallback emails<br>
            (semicolon, comma, pipe, or <code> * </code> separated).<br><br>
            <strong style="color:#16a34a">What's fixed in v2:</strong><br>
            ✓ SMTP probe: STARTTLS + port 587 fallback<br>
            ✓ MX records sorted by priority before probing<br>
            ✓ SMTP 550 = definitive Not Deliverable<br>
            ✓ SMTP blocked → DNS signals decide (no false fails)<br>
            ✓ CSV export alongside XLSX<br>
            ✓ Exports built once post-run, not every rerun<br>
            ✓ Queue cached during validation (no reparse lag)<br>
            ✓ Live table shows last 10 rows only during validation
        </div>
    </div>""", unsafe_allow_html=True)
